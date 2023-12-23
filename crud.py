import openpyxl

def create_table():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "DataMahasiswa"
    ws.append(["ID", "Nama", "Umur"])
    wb.save("data_mahasiswa.xlsx")
    
def insert_data(nama, umur):
    wb = openpyxl.load_workbook("data_mahasiswa.xlsx")
    ws = wb["DataMahasiswa"]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)
    (ID, name, age) = data[-1]
    ws.append([ID+1, nama, umur])
    wb.save("data_mahasiswa.xlsx")
    
def read_data():
    wb = openpyxl.load_workbook("data_mahasiswa.xlsx")
    ws = wb["DataMahasiswa"]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data
    
def update_data(ID, new_nama, new_umur):
    wb = openpyxl.load_workbook("data_mahasiswa.xlsx")
    ws = wb["DataMahasiswa"]
    
    found_row_index = None

    for row_index, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True), start=2):
        (ID_cari, nama, umur) = row
        if ID + 1 == ID_cari:
            found_row_index = row_index
            break

    if found_row_index is not None:
        ws.cell(row=found_row_index, column=2, value=new_nama)
        ws.cell(row=found_row_index, column=3, value=new_umur)
        wb.save("data_mahasiswa.xlsx")
    else:
        print(f"Data dengan ID {ID} tidak ditemukan.")
            
def delete_data(ID):
    wb = openpyxl.load_workbook("data_mahasiswa.xlsx")
    ws = wb["DataMahasiswa"]
    
    found_row_index = None

    for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        (ID_cari, nama, umur) = row
        if ID+1 == ID_cari:
            found_row_index = row_index
            break

    if found_row_index is not None:
        ws.delete_rows(found_row_index)
        wb.save("data_mahasiswa.xlsx")
    else:
        print(f"Data dengan ID {ID+1} tidak ditemukan.")